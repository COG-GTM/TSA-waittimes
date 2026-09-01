#!/usr/bin/env python3
"""Fetch FAA CY2024 commercial-service enplanements and write the static data file.

Source page: https://www.faa.gov/airports/planning_capacity/passenger_allcargo_stats/passenger
Source file: https://www.faa.gov/airports/planning_capacity/passenger_allcargo_stats/passenger/arp-cy2024-commercial-service-enplanements.xlsx

Re-run annually when FAA publishes the next CY file.
"""
from __future__ import annotations

import argparse
import asyncio
import io
import json
import sys
from datetime import UTC, datetime
from pathlib import Path

from curl_cffi.requests import AsyncSession
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.enplanements import parse_enplanement_rows

SOURCE_PAGE = "https://www.faa.gov/airports/planning_capacity/passenger_allcargo_stats/passenger"
SOURCE_URL = "https://www.faa.gov/airports/planning_capacity/passenger_allcargo_stats/passenger/arp-cy2024-commercial-service-enplanements.xlsx"
SOURCE_NAME = "FAA CY2024 Commercial Service Enplanements (all commercial service airports)"
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "data" / "enplanements.json"


async def download() -> bytes:
    async with AsyncSession(impersonate="chrome") as session:
        response = await session.get(SOURCE_URL, timeout=30)
        response.raise_for_status()
        return response.content


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", type=Path, help="parse a local XLSX copy instead of downloading")
    args = parser.parse_args()
    workbook_bytes = args.file.read_bytes() if args.file else asyncio.run(download())
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    year, records = parse_enplanement_rows(worksheet.iter_rows(values_only=True))
    airports = sorted(records, key=lambda record: int(record["rank"]))
    output = {
        "year": year,
        "source_name": SOURCE_NAME,
        "source_page": SOURCE_PAGE,
        "source_url": SOURCE_URL,
        "retrieved_at": datetime.now(tz=UTC).isoformat(),
        "airports": airports,
    }
    with OUTPUT_PATH.open("w", encoding="utf-8") as file:
        json.dump(output, file, indent=2)
        file.write("\n")
    print(f"Year: {year}")
    print(f"Airports: {len(airports)}")
    print("Top 5:")
    for airport in airports[:5]:
        print(f"  {airport['rank']}. {airport['locid']}: {airport['enplanements']:,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
